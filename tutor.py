import static
from openpyxl import load_workbook

WEEKS = 13
HOURS = 10
DAYS = 5
START_HOUR = 9
AVAIL_PATH = "C:\\Users\\Liam\\Google Drive\\Personal\\Employment\\Tutoring\\Tutoring Availability.xlsx"

ROW_OFFSET = 2
COLUMN_OFFSET = 2

def create_tabs(num_weeks):
    """
    Create tabs containing availability tables for the given number of weeks. 

    Arguments:
        num_weeks (int): number of weeks for which tabs will be created

    Returns (str): HTML page contents
    """
    tab_names = []
    page = static.get_template("tutoring")
    tabs = ""
    for i in range(0, num_weeks):
        week = f"Week {i + 1}"
        table = create_availability_table(week)
        tabs += f"""<div id="{i + 1}" class="tabcontent">
        {table}
        </div>\n
        """
        tab_names.append(i + 1)
    page = page.replace("{ TAB_CONTENT }", tabs)

    tab_links = ""
    for i, name in enumerate(tab_names):
        if i == 0:
            tab_links += f'<button class="tablinks" id="defaultOpen" onclick="openTab(event, \'{name}\')">{name}</button>\n'
        else:
            tab_links += f'<button class="tablinks" onclick="openTab(event, \'{name}\')">{name}</button>\n'
    page = page.replace("{ TAB_LINKS }", tab_links)
    
    return page


def create_availability_table(week):
    """
    Create availability table for the given week.

    Arguments:
        week (str): name of week
    """
    table = """<table>
    <tr>
        <td class="empty"></td>
        <th>Mon</th>
        <th>Tue</th>
        <th>Wed</th>
        <th>Thu</th>
        <th>Fri</th>
    </tr>\n
    """
    availability = load_workbook(AVAIL_PATH)[week]
    for i in range(0, HOURS):
        table += f"<tr>\n<td>{i + START_HOUR}:00</td>\n"
        for j in range(0, DAYS):
            cell = availability.cell(i + ROW_OFFSET, j + COLUMN_OFFSET)
            cell_class = "unavailable" if cell.value else "available"
            table += f"<td class={cell_class}></td>\n"
        table += "</tr>\n"
    table += "</table>"
    return table


if __name__ == "__main__":
    page = create_tabs(WEEKS)
    with open("tutoring", "w") as f:
        f.write(page)
